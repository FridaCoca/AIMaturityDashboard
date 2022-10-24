from pathlib import PurePath

from openpyxl import Workbook

from network.sharepoint_remote_data_source import SharePointRemoteDataSource


def set_file_ext(file_name, export_type):
    if export_type == 'Excel':
        file_name_with_ext = '.'.join([file_name, '.xlsx'])
    elif export_type == 'CSV':
        file_name_with_ext = '.'.join([file_name, '.csv'])
    else:
        file_name_with_ext = file_name
    return file_name_with_ext


def download_list(list_name, export_type, dir_path, file_name):
    sp_list = SharePointRemoteDataSource().get_list(list_name)
    if export_type == 'Excel':
        save_to_excel(sp_list, dir_path, file_name)
    else:
        print('Export type is not a value type')


def save_to_excel(list_items, dir_path, file_name):
    dir_file_path = PurePath(dir_path, file_name)
    wb = Workbook()
    ws = wb.active
    if len(list_items) == 0:
        print("no list found")
        return
    # list of header name from SharePoint List
    header = list_items[0].properties.keys()
    # write headers on first row
    for idx, name in enumerate(header):
        ws.cell(row=1, column=idx + 1, value=name)
    # write line items starting on second row
    row = 2
    for dict_obj in list_items:
        for idx, item in enumerate(dict_obj.properties.items()):
            ws.cell(row=row, column=idx + 1, value=item[1])
        row += 1
    wb.save(dir_file_path)
